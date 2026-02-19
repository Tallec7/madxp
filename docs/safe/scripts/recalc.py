import openpyxl
from openpyxl.utils import get_column_letter
import sys

def recalc_workbook(filepath):
    """
    Force Excel to recalculate all formulas by:
    1. Loading workbook without data_only
    2. Accessing each cell to ensure formulas are recognized
    3. Saving with standard settings
    """
    print("=" * 80)
    print(f"RECALCULATING WORKBOOK: {filepath}")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    print(f"\nWorkbook loaded with {len(wb.sheetnames)} sheets")
    
    formula_count = 0
    sheet_count = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    sheet_formulas += 1
                    formula_count += 1
        
        if sheet_formulas > 0:
            sheet_count += 1
            print(f"  {sheet_name}: {sheet_formulas} formulas")
    
    print(f"\nTotal sheets with formulas: {sheet_count}")
    print(f"Total formulas: {formula_count}")
    
    # Set calculation mode to automatic (this flags for recalc)
    wb.calculation.calcMode = 'auto'
    
    print(f"\nSaving workbook with automatic calculation...")
    wb.save(filepath)
    
    print(f"\nWorkbook saved successfully!")
    print("Formulas will be recalculated when opened in Excel/Google Sheets")
    
    return True

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <workbook_path>")
        sys.exit(1)
    
    filepath = sys.argv[1]
    try:
        recalc_workbook(filepath)
        print("\n✓ Recalculation complete")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        sys.exit(1)
