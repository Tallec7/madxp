#!/usr/bin/env python3
"""
export-to-excel.py — Génère NEOPRO_SAFe_Portfolio.xlsx depuis les fichiers .md

Usage:
    python docs/safe/scripts/export-to-excel.py
    python docs/safe/scripts/export-to-excel.py --output /path/to/output.xlsx

Source de vérité : docs/safe/*.md (git)
Résultat : docs/safe/NEOPRO_SAFe_Portfolio.xlsx (livrable Excel)

Structure Excel (13 onglets) :
    1. Dashboard            — KPIs, status, roadmap, velocity (formules cross-sheet)
    2. Glossaire            — Termes SAFe pour les non-techniques
    3. Vision & OKR         — Vision, 4 Thèmes, 5 Objectifs, 38 Key Results, KPIs
    4. Value Streams        — OVS-1, OVS-2, DVS-1 détaillé, 12 outils Neopro
    5. Epics & LBC          — 21 Epics avec WSJF, SUMIF Cost dynamique
    6. Features & US        — 35 Features groupées par PI, COUNTIF/SUMIF → User Stories
    7. PI Objectives        — Objectifs PI-1/2 + aspirationnel, résumé BV, prédictabilité
    8. Sprint Tracker       — S1-S12 (PI-1 + PI-2), résumé, capacité équipe
    9. ROAM                 — 8 risques avec action/atténuation, résumé COUNTIF
   10. Flow Metrics         — Métriques par VS + allocation avec barres REPT
   11. Implemented Backlog  — 178+ features, statistiques produit, domain summary COUNTIF
   12. User Stories         — Backlog US détaillé (headers, lié à Features)
   13. _ChartData           — Données pour graphiques Excel
"""

import re
import os
import sys
import csv
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# === Paths ===
SCRIPT_DIR = Path(__file__).parent
SAFE_DIR = SCRIPT_DIR.parent
DEFAULT_OUTPUT = SAFE_DIR / "NEOPRO_SAFe_Portfolio.xlsx"

# === Style constants ===
BLUE_DARK = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
BLUE_MED = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
GREEN_LIGHT = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
ORANGE_LIGHT = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
RED_LIGHT = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
BLUE_LIGHT = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
PURPLE_LIGHT = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
GREY_LIGHT = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=12)
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
LINK_FONT = Font(color="1565C0", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def read_md(filename):
    """Read a markdown file from docs/safe/."""
    path = SAFE_DIR / filename
    if not path.exists():
        print(f"  Warning: {filename} not found, skipping")
        return ""
    return path.read_text(encoding="utf-8")


def set_header_row(ws, row, headers, fill=None):
    """Write a bold header row with optional fill."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if fill:
            cell.fill = fill


def auto_width(ws, min_width=10, max_width=50):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def write_row(ws, row, data, start_col=1):
    """Write a row of data with borders."""
    for j, val in enumerate(data, start_col):
        cell = ws.cell(row=row, column=j, value=val)
        cell.border = THIN_BORDER
    return row + 1


# ============================================================
# Sheet builders — aligned with NEOPRO_SAFe_PortfoliovTravail.xlsx
# ============================================================

def build_dashboard(wb):
    """Sheet 1: Dashboard — KPIs, status, roadmap, velocity, navigation."""
    ws = wb.active
    ws.title = "Dashboard"

    # R1: Title
    ws["A1"] = "NEOPRO SAFe Portfolio Dashboard"
    ws["A1"].font = TITLE_FONT

    # R3: PORTFOLIO OVERVIEW
    ws["A3"] = "PORTFOLIO OVERVIEW"
    ws["A3"].font = HEADER_FONT

    # R4: Counters (formulas)
    ws["A4"] = "Epics"
    ws["B4"] = "=COUNTA('Epics & LBC'!A4:A24)-COUNTBLANK('Epics & LBC'!A4:A24)"
    ws["C4"] = "Features"
    ws["D4"] = "=COUNTA('Features & US'!B5:B41)-COUNTBLANK('Features & US'!B5:B41)"
    ws["E4"] = "User Stories"
    ws["F4"] = "=SUM('Features & US'!D5:D41)"
    ws["G4"] = "Total SP"
    ws["H4"] = "=SUM('Features & US'!E5:E41)"

    # R5: SP by PI
    ws["A5"] = "SP PI-1"
    ws["B5"] = "=SUMIF('Features & US'!I:I,\"PI-1*\",'Features & US'!E:E)"
    ws["C5"] = "SP PI-2"
    ws["D5"] = "=SUMIF('Features & US'!I:I,\"PI-2*\",'Features & US'!E:E)"
    ws["E5"] = "SP PI-3"
    ws["F5"] = "=SUMIF('Features & US'!I:I,\"PI-3*\",'Features & US'!E:E)"
    ws["G5"] = "ARR (K\u20ac)"
    ws["H5"] = 350

    # R7: Status sections
    ws["A7"] = "EPICS STATUS"
    ws["A7"].font = HEADER_FONT
    ws["C7"] = "FEATURES STATUS"
    ws["C7"].font = HEADER_FONT
    ws["E7"] = "RISKS (ROAM)"
    ws["E7"].font = HEADER_FONT
    ws["G7"] = "FLOW DISTRIBUTION"
    ws["G7"].font = HEADER_FONT

    # R8-R11: Status counts
    ws["A8"] = "En cours"
    ws["B8"] = "=COUNTIF('Epics & LBC'!G:G,\"En cours\")"
    ws["C8"] = "Termin\u00e9"
    ws["D8"] = "=COUNTIF('Features & US'!G5:G41,\"Termin\u00e9\")"
    ws["E8"] = "Open"
    ws["F8"] = '=COUNTIF(ROAM!C:C,"Owned")+COUNTIF(ROAM!C:C,"Accepted")'  # vTravail uses French but formulas ref English-mapped
    ws["G8"] = "Mon\u00e9tisation"
    ws["H8"] = 0.4

    ws["A9"] = "Termin\u00e9"
    ws["B9"] = "=COUNTIF('Epics & LBC'!G:G,\"Termin\u00e9\")"
    ws["C9"] = "Backlog"
    ws["D9"] = "=COUNTIF('Features & US'!G5:G41,\"Backlog\")"
    ws["E9"] = "Resolved"
    ws["F9"] = '=COUNTIF(ROAM!C:C,"Resolved")+COUNTIF(ROAM!C:C,"Mitigated")'
    ws["G9"] = "Exp. Match"
    ws["H9"] = 0.25

    ws["A10"] = "Backlog"
    ws["B10"] = "=COUNTIF('Epics & LBC'!G:G,\"Backlog\")"
    ws["C10"] = "En cours"
    ws["D10"] = "=COUNTIF('Features & US'!G5:G41,\"En cours\")"
    ws["E10"] = "Owned"
    ws["F10"] = '=COUNTIF(ROAM!C:C,"Owned")'
    ws["G10"] = "Acquisition"
    ws["H10"] = 0.2

    ws["A11"] = "Total"
    ws["B11"] = "=B8+B9+B10"
    ws["C11"] = "Total"
    ws["D11"] = "=D8+D9+D10"
    ws["E11"] = "Excellence Ops"
    ws["G11"] = "Excellence Ops"
    ws["H11"] = 0.15

    # R13-R17: Roadmap
    ws["A13"] = "ROADMAP SUMMARY"
    ws["A13"].font = HEADER_FONT
    set_header_row(ws, 14, ["PI", "P\u00e9riode", "Initiatives cl\u00e9s", "", "", "", "Planned SP", "Status"])
    write_row(ws, 15, ["PI-1", "F\u00e9v-Mar", "Sponsors + Analytics + Onboarding + R\u00e9silience WiFi", "", "", "",
                        "=SUMIF('Features & US'!I:I,\"PI-1*\",'Features & US'!E:E)", "\U0001f7e2 Active"])
    write_row(ws, 16, ["PI-2", "Avr-Mai", "R\u00e9gie + Motion + Email + Score API F\u00e9d\u00e9rations", "", "", "",
                        "=SUMIF('Features & US'!I:I,\"PI-2*\",'Features & US'!E:E)", "\U0001f7e1 Planned"])
    write_row(ws, 17, ["PI-3", "Jun-Jul", "Multi-\u00e9crans + Marque blanche + Billetterie + ML + OAuth", "", "", "",
                        "=SUMIF('Features & US'!I:I,\"PI-3*\",'Features & US'!E:E)", "\u26aa Backlog"])

    # R19-R21: Sprint Velocity
    ws["A19"] = "SPRINT VELOCITY (PI-1)"
    ws["A19"].font = HEADER_FONT
    write_row(ws, 20, ["Sprint", "S1", "S2", "S3", "S4", "S5", "S6", "Avg"])
    write_row(ws, 21, ["Actual SP",
                        "='Sprint Tracker'!D6",
                        "='Sprint Tracker'!D7",
                        "='Sprint Tracker'!D8",
                        "='Sprint Tracker'!D9",
                        "=IF('Sprint Tracker'!D10=\"\",\"-\",'Sprint Tracker'!D10)",
                        "=IF('Sprint Tracker'!D11=\"\",\"-\",'Sprint Tracker'!D11)",
                        "=IFERROR(AVERAGE('Sprint Tracker'!D6:D9),0)"])

    # R23-R25: Navigation
    ws["A23"] = "NAVIGATION"
    ws["A23"].font = HEADER_FONT
    write_row(ws, 24, ["Vision & OKR", "Value Streams", "Epics & LBC", "Features & US",
                        "PI Objectives", "Sprint Tracker", "ROAM", "Flow Metrics"])
    write_row(ws, 25, ["Glossaire", "Impl. Backlog"])

    # R27: Charts placeholder
    ws["A27"] = "CHARTS"
    ws["A27"].font = HEADER_FONT

    auto_width(ws)
    return ws


def build_glossaire(wb):
    """Sheet 2: Glossaire SAFe pour Gabin."""
    ws = wb.create_sheet("Glossaire")
    ws["A1"] = "Glossaire SAFe pour Gabin"
    ws["A1"].font = TITLE_FONT

    terms = [
        ("Epic", "Grande initiative couvrant plusieurs Features, li\u00e9e \u00e0 une strat\u00e9gie business"),
        ("Feature", "Bloc fonctionnel apportant valeur, d\u00e9coup\u00e9 en User Stories"),
        ("User Story", "R\u00e9cit utilisateur livrable en 1-2 sprints (taille ~5 SP)"),
        ("Story Point (SP)", "Unit\u00e9 de complexit\u00e9 relative (Fibonacci: 1,2,3,5,8...)"),
        ("Sprint", "It\u00e9ration 2 semaines de travail continu"),
        ("PI (Program Increment)", "Trim de 3 mois = 6 sprints align\u00e9s"),
        ("Value Stream", "Flux continu de cr\u00e9ation de valeur (OVS ou DVS)"),
        ("WSJF (Weighted Shortest Job First)", "Priorisation: (User-Biz-Time-RR) / Job Size"),
        ("MoSCoW", "Must (obligatoire), Should (important), Could (sympa), Won't (exclu)"),
        ("Lean Business Case", "Justification business simplifi\u00e9e (co\u00fbt vs b\u00e9n\u00e9fice)"),
        ("Crit\u00e8res d'acceptation", "Conditions pour que la US soit \"Done\" (d\u00e9finition"),
        ("Backlog", "Pile ordonn\u00e9e de Features/US pr\u00eates ou futures"),
        ("Velocity", "SP livr\u00e9s en moyenne par sprint (mesure de cadence)"),
        ("Carry-over", "US commenc\u00e9e mais non finie avant fin de sprint"),
        ("Business Value (BV)", "Score 1-10 indiquant impact client/revenu (PI Objective)"),
        ("Predictability", "% d'objectives commis atteints dans le PI"),
        ("ROAM", "Risk log: Resolved, Owned, Accepted, Mitigated"),
        ("Flow Metrics", "WIP, Throughput, Cycle Time pour chaque Value Stream"),
        ("WIP (Work In Progress)", "Limite de t\u00e2ches simultan\u00e9es pour fluider"),
        ("CFD (Cumulative Flow Diagram)", "Graphe d'accumulation des status (backlog \u2192 done)"),
        ("DVS (Dedicated Value Stream)", "VS interne Neopro (plateforme produit)"),
        ("OVS (Outcome Value Stream)", "VS client (club \u00e0 \u00e9cran, sponsor \u00e0 impression)"),
        ("I&A (Inspect & Adapt)", "R\u00e9trospective SAFe collective tous les PIs"),
    ]

    set_header_row(ws, 2, ["Terme", "D\u00e9finition"], fill=BLUE_LIGHT)
    for i, (term, definition) in enumerate(terms, 3):
        ws.cell(row=i, column=1, value=term).border = THIN_BORDER
        ws.cell(row=i, column=2, value=definition).border = THIN_BORDER
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70


def build_vision_okr(wb):
    """Sheet 3: Vision & OKR — 5 Objectifs, 38 Key Results, KPIs actuels (64 rows)."""
    ws = wb.create_sheet("Vision & OKR")
    ws["A1"] = "NEOPRO \u2014 Vision, Th\u00e8mes Strat\u00e9giques & OKR 2026"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Vision Produit"
    ws["A3"].font = HEADER_FONT
    ws["A5"] = "Devenir le partenaire de r\u00e9f\u00e9rence des clubs sportifs amateurs en France pour promouvoir leur identit\u00e9, engager leurs communaut\u00e9s et mon\u00e9tiser leur audience via une solution TV interactive cl\u00e9-en-main."

    # Thèmes Stratégiques
    ws["A7"] = "4 Th\u00e8mes Strat\u00e9giques"
    ws["A7"].font = HEADER_FONT
    set_header_row(ws, 8, ["ID", "Th\u00e8me", "Description", "OVS li\u00e9", "Horizon", "Poids budget"], fill=RED_LIGHT)
    themes = [
        ("TS-1", "Mon\u00e9tisation Sponsors", "Maximiser les revenus sponsors par club via le ciblage local, les preuves de diffusion et le self-service", "OVS-2", "PI-1 \u2192 PI-3", "35%"),
        ("TS-2", "Exp\u00e9rience Match Live", "Offrir la meilleure exp\u00e9rience match jour (overlay multi-sport, timer, recording)", "OVS-1", "PI-1 \u2192 PI-2", "25%"),
        ("TS-3", "Acquisition & Croissance", "Acc\u00e9l\u00e9rer l'onboarding clubs (time-to-value <1j), r\u00e9duire le churn (<5%/an), atteindre 100 clubs", "OVS-1 + OVS-2", "PI-2 \u2192 PI-3", "20%"),
        ("TS-4", "Excellence Op\u00e9rationnelle", "Garantir 99.5% uptime, <2h MTTR, monitoring proactif, OTA fiable, r\u00e9seau r\u00e9silient", "DVS-1", "Continu", "20%"),
    ]
    for i, row_data in enumerate(themes, 9):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val).border = THIN_BORDER

    # OKR 2026 — 5 Objectifs, 38 Key Results
    ws["A14"] = "OKR 2026 \u2014 Objectifs & Key Results"
    ws["A14"].font = HEADER_FONT
    set_header_row(ws, 15, ["Obj #", "Objectif", "KR #", "Key Result", "Th\u00e8me(s)", "Target"], fill=BLUE_LIGHT)

    okrs = [
        ("O1", "Faire de NEOPRO le partenaire de r\u00e9f\u00e9rence des clubs amateurs en Bretagne/Pays de Loire", "1.1", "3 clubs beta convertis en payants", "TS-3", "Q2 2026"),
        ("", "", "1.2", "15 clubs payants actifs", "TS-3", "D\u00e9c 2026"),
        ("", "", "1.3", "Taux de renouvellement >85%", "TS-3", "D\u00e9c 2026"),
        ("", "", "1.4", "NPS clubs \u22658/10", "TS-2, TS-3", "Q4 2026"),
        ("", "", "1.5", "2 partenariats ligues r\u00e9gionales sign\u00e9s", "TS-3", "Q3 2026"),
        ("", "", "1.6", "3 cas ROI sponsors document\u00e9s", "TS-1", "Q3 2026"),
        ("", "", "1.7", "6 annonceurs r\u00e9gionaux actifs", "TS-1", "Q4 2026"),
        ("", "", "1.8", "ARPU annonceur >200\u20ac/mois", "TS-1", "Q4 2026"),
        ("O2", "Cr\u00e9er une exp\u00e9rience match que les clubs ne peuvent plus quitter", "2.1", "Taux d'utilisation 100% des matchs \u00e0 domicile", "TS-2", "Q4 2026"),
        ("", "", "2.2", "Latence t\u00e9l\u00e9commande <200ms local / <2s cloud", "TS-2", "Q1 2026"),
        ("", "", "2.3", "6 sports support\u00e9s en overlay", "TS-2", "\u2705 Done"),
        ("", "", "2.4", "Score CSAT b\u00e9n\u00e9vole >4.5/5", "TS-2", "Q3 2026"),
        ("", "", "2.5", "Taux activation >80% (1er match dans les 7 jours)", "TS-2, TS-3", "Continu"),
        ("", "", "2.6", "0 incident bloquant pendant un match", "TS-2, TS-4", "Continu"),
        ("", "", "2.7", "Temps d'installation <4h par club", "TS-2, TS-4", "Q2 2026"),
        ("O3", "B\u00e2tir un r\u00e9seau de clubs qui tire la croissance", "3.1", "30% des nouveaux clubs via referral", "TS-3", "Q4 2026"),
        ("", "", "3.2", "10 t\u00e9moignages vid\u00e9o clubs", "TS-3", "Q3 2026"),
        ("", "", "3.3", "2 clubs ambassadeurs officiels par r\u00e9gion", "TS-3", "Q4 2026"),
        ("", "", "3.4", "15 prospects qualifi\u00e9s via r\u00e9seau", "TS-3", "Q3 2026"),
        ("", "", "3.5", "Pr\u00e9sence dans 2 m\u00e9dias sportifs r\u00e9gionaux", "TS-3", "Q3 2026"),
        ("", "", "3.6", "1 partenariat tournoi majeur", "TS-3", "Q4 2026"),
        ("", "", "3.7", "NPS \u226550", "TS-3", "Q4 2026"),
        ("", "", "3.8", "25 clubs onboard\u00e9s en Q1-Q2", "TS-3", "Juin 2026"),
        ("O4", "Garantir une infrastructure fiable \u00e0 la hauteur de l'ambition", "4.1", "Uptime API 99.5%", "TS-4", "Continu"),
        ("", "", "4.2", "MTTR incident <2h", "TS-4", "Continu"),
        ("", "", "4.3", "0 perte de donn\u00e9es vid\u00e9o/config (checksum SHA-256)", "TS-4", "Continu"),
        ("", "", "4.4", "OTA success rate >95%", "TS-4", "Q2 2026"),
        ("", "", "4.5", "Temps de r\u00e9solution support <24h", "TS-4", "Continu"),
        ("", "", "4.6", "Monitoring proactif en place (alertes automatiques)", "TS-4", "Q2 2026"),
        ("", "", "4.7", "Proc\u00e9dure d'installation document\u00e9e reproductible sans Gabin", "TS-4", "Q2 2026"),
        ("O5", "Construire un produit scalable sans dette technique", "5.1", "API partenaires <100ms p95", "TS-4", "Q4 2026"),
        ("", "", "5.2", "Tests coverage >75% (actuellement ~2387 tests)", "TS-4", "Q3 2026"),
        ("", "", "5.3", "Pipeline CI <5min", "TS-4", "Q2 2026"),
        ("", "", "5.4", "0 vuln\u00e9rabilit\u00e9 critique non patch\u00e9e sous 48h", "TS-4", "Continu"),
        ("", "", "5.5", "Documentation technique \u00e0 jour pour 100% des endpoints", "TS-4", "Q3 2026"),
        ("", "", "5.6", "Temps d'onboarding dev externe <1 jour", "TS-4", "Q3 2026"),
        ("", "", "5.7", "Architecture supportant 200 clubs sans refonte", "TS-4", "Q4 2026"),
        ("", "", "5.8", "Roadmap produit Q3-Q4 2026 d\u00e9finie et prioris\u00e9e", "TS-4", "Q2 2026"),
    ]
    for i, row_data in enumerate(okrs, 16):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val).border = THIN_BORDER

    r = 16 + len(okrs)
    ws.cell(row=r, column=1, value=f"Total: 5 Objectifs, {len(okrs)} Key Results")

    # KPI Actuels
    r += 2
    ws.cell(row=r, column=1, value="KPI Actuels (F\u00e9v 2026)").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["KPI", "Valeur actuelle", "Target", "Statut", "Source"], fill=GREEN_LIGHT)
    kpis = [
        ("Clubs actifs (Pi connect\u00e9s)", "4", "100", "\U0001f7e1 En cours", "Dashboard fleet"),
        ("Annonceurs actifs", "0", "10", "\U0001f7e1 Early stage", "DB advertisers"),
        ("Tests totaux", "2 235", "2 500", "\U0001f7e2 On track", "npm run test:*"),
        ("Uptime API", "99.2%", "99.5%", "\U0001f7e1 Proche", "Grafana"),
        ("Versions livr\u00e9es", "30+ (v3.47\u2192v3.60)", "Continu", "\U0001f7e2 Actif", "Git tags"),
        ("Features impl\u00e9ment\u00e9es", "184+", "-", "\U0001f7e2 Complet", "IMPLEMENTED-BACKLOG"),
    ]
    for i, row_data in enumerate(kpis, r + 1):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val).border = THIN_BORDER

    auto_width(ws)


def build_value_streams(wb):
    """Sheet 4: Value Streams — OVS-1/2 détail, DVS-1 détail, 12 outils."""
    ws = wb.create_sheet("Value Streams")
    ws["A1"] = "NEOPRO \u2014 Value Streams & Flux de Valeur"
    ws["A1"].font = TITLE_FONT

    # OVS Summary
    ws["A3"] = "Outcome Value Streams (OVS)"
    ws["A3"].font = HEADER_FONT
    set_header_row(ws, 4, ["ID", "Value Stream", "Trigger", "Value Delivered", "Key Steps", "Lead Time", "Bottleneck"], fill=GREEN_LIGHT)
    vs_summary = [
        ("OVS-1", "Club to Screen", "Club signe contrat", "Vid\u00e9os diffus\u00e9es sur TV dans le gymnase",
         "Onboarding \u2192 Config \u2192 Upload \u2192 Deploy \u2192 Diffusion", "< 1 jour (target)", "Installation physique Pi"),
        ("OVS-2", "Sponsor to Impression", "Annonceur signe contrat", "Impressions pub mesur\u00e9es et certifi\u00e9es",
         "Signup \u2192 Upload pub \u2192 Association sites \u2192 Diffusion \u2192 Report", "< 2 jours", "Association manuelle site-sponsor"),
    ]
    for i, row_data in enumerate(vs_summary, 5):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)

    # OVS-1 detail
    r = 8
    ws.cell(row=r, column=1, value="OVS-1 \u2014 Club to Screen (d\u00e9tail)").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["#", "Step", "Outil", "Owner", "Dur\u00e9e", "Input", "Output"], fill=GREEN_LIGHT)
    ovs1_steps = [
        (1, "Vente & signature", "CRM (externe)", "Commercial", "1-2 semaines", "Lead club", "Contrat sign\u00e9 + abonnement"),
        (2, "Installation Pi dans le club", "Toolbox D\u00e9ploiement", "Technicien", "1-2 heures", "Pi + TV + SD card", "Pi connect\u00e9 au Dashboard"),
        (3, "Configuration club", "Dashboard Admin", "Op\u00e9rateur", "15 min", "Logo, cat\u00e9gories, sponsors", "Config d\u00e9ploy\u00e9e sur Pi"),
        (4, "Upload vid\u00e9os contenu", "Dashboard Admin", "Op\u00e9rateur", "10 min", "Fichiers vid\u00e9o", "Vid\u00e9os sur FTP + checksum"),
        (5, "D\u00e9ploiement vers Pi", "Central Server \u2192 Sync Agent", "Automatis\u00e9", "2-5 min", "Commande deploy", "Vid\u00e9os sur disque Pi"),
        (6, "Diffusion sur TV", "TV Player", "Automatis\u00e9", "Imm\u00e9diat", "Config + vid\u00e9os", "Boucle vid\u00e9o diffus\u00e9e"),
        (7, "Match day (b\u00e9n\u00e9vole)", "T\u00e9l\u00e9commande", "Staff club", "Dur\u00e9e match", "QR code / URL", "Score live + phases + recording"),
        (8, "Analytics & rapports", "Dashboard + PDF", "Automatis\u00e9", "Quotidien", "Impressions brutes", "Stats club + sponsor"),
    ]
    for i, row_data in enumerate(ovs1_steps, r + 1):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val).border = THIN_BORDER

    # OVS-2 detail
    r = r + len(ovs1_steps) + 2
    ws.cell(row=r, column=1, value="OVS-2 \u2014 Sponsor to Impression (d\u00e9tail)").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["#", "Step", "Outil", "Owner", "Dur\u00e9e", "Input", "Output"], fill=ORANGE_LIGHT)
    ovs2_steps = [
        (1, "Inscription annonceur", "Portail Annonceur / Magic Link", "Annonceur", "5 min", "Email + vid\u00e9o pub", "Compte cr\u00e9\u00e9 + vid\u00e9o upload\u00e9e"),
        (2, "Association site-sponsor", "Dashboard Admin", "Op\u00e9rateur", "2 min", "Sponsor + sites cibles", "Mapping sponsor\u2192sites"),
        (3, "Sync sponsors vers Pi", "Central Server \u2192 Sync Agent", "Automatis\u00e9", "2-5 min", "Config updated", "Sponsors sur disque Pi"),
        (4, "Diffusion pub sur TV", "TV Player", "Automatis\u00e9", "Imm\u00e9diat", "Boucle vid\u00e9o avec pubs", "Impressions track\u00e9es"),
        (5, "Tracking impressions", "TV Player \u2192 Sync Agent \u2192 API", "Automatis\u00e9", "Batch quotidien", "Buffer local Pi", "Impressions en DB"),
        (6, "Proof of broadcast", "Central Server", "Automatis\u00e9", "Temps r\u00e9el", "Screenshot + timestamp", "Preuve certifi\u00e9e SHA-256"),
        (7, "Rapport annonceur", "Portail Annonceur / PDF", "Annonceur", "Self-service", "Stats agr\u00e9g\u00e9es", "PDF/Excel export"),
    ]
    for i, row_data in enumerate(ovs2_steps, r + 1):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val).border = THIN_BORDER

    # DVS-1 detail
    r = r + len(ovs2_steps) + 2
    ws.cell(row=r, column=1, value="DVS-1 \u2014 Neopro Platform (Development Value Stream)").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["#", "Step", "Owner", "Outils", "M\u00e9triques", "Cadence"], fill=PURPLE_LIGHT)
    dvs1_steps = [
        (1, "Product Discovery & Backlog", "PO (Gwenvael)", "SAFe Framework, Notion", "WSJF score, backlog depth", "Continu"),
        (2, "Architecture & Design", "Tech Lead", "ADR, TypeScript strict", "Design reviews pass\u00e9s", "Par feature"),
        (3, "Development (Sprint)", "Engineering", "Git, ESLint, Repository Pattern", "V\u00e9locit\u00e9 SP/sprint, WIP", "Sprint 2 semaines"),
        (4, "Testing & QA", "Engineering", "Jest, Karma, Playwright, Smoke", "2 235 tests, coverage", "CI automatique"),
        (5, "Release & OTA", "DevOps", "Git tags, Railway, FTP, Canary", "OTA success rate, rollback", "~2 releases/semaine"),
        (6, "Monitoring & Feedback", "Support", "Grafana, Prometheus, Slack", "Uptime, MTTR, NPS", "Continu"),
    ]
    for i, row_data in enumerate(dvs1_steps, r + 1):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val).border = THIN_BORDER

    # 12 Outils Neopro
    r = r + len(dvs1_steps) + 2
    ws.cell(row=r, column=1, value="Cartographie 12 Outils Neopro").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["#", "Outil", "Type", "Public", "Protocole", "VS li\u00e9", "Features cl\u00e9s"], fill=GREY_LIGHT)
    tools = [
        (1, "Dashboard Admin", "Produit cloud", "Admin, Op\u00e9rateurs", "HTTP + WebSocket", "OVS-1, OVS-2", "Fleet mgmt, deploy, analytics, alerting, OTA"),
        (2, "Portail Annonceur", "Produit cloud", "Annonceurs", "HTTP + JWT", "OVS-2", "Upload pub, stats impressions, export PDF"),
        (3, "Portail Agence", "Produit cloud", "Agences pub", "HTTP + JWT", "OVS-2", "Multi-annonceurs, stats agr\u00e9g\u00e9es"),
        (4, "T\u00e9l\u00e9commande", "Produit cloud+edge", "Staff club", "Socket.IO", "OVS-1", "Score, phases, timer, recording, vid\u00e9os"),
        (5, "TV Player", "Produit edge", "Spectateurs", "Socket.IO local", "OVS-1, OVS-2", "Double-buffer, overlay 6 sports, watermark"),
        (6, "Admin Panel Pi", "Produit edge", "Staff + Tech", "HTTP local", "OVS-1", "Club/Tech mode, upload local, debug"),
        (7, "Central Server API", "Infra cloud", "Machines", "REST + WebSocket", "DVS-1", "21 repos, 9 handlers, rate limiting"),
        (8, "Sync Agent", "Infra edge", "Syst\u00e8me", "Socket.IO", "OVS-1, OVS-2", "Config sync, video download, analytics batch"),
        (9, "Pi Server Socket.IO", "Infra edge", "T\u00e9l\u00e9commande, TV", "Socket.IO local", "OVS-1", "18 \u00e9v\u00e9nements relay, score, phases"),
        (10, "Watchdogs (4)", "Infra edge", "Syst\u00e8me", "Bash + systemd", "DVS-1", "Kiosk, sync-agent, network, hotspot"),
        (11, "Monitoring", "Ops", "Support", "Prometheus + Grafana", "DVS-1", "30+ m\u00e9triques, 18 seuils, 3 dashboards"),
        (12, "Toolbox D\u00e9ploiement", "Ops", "Techniciens", "Bash + SSH", "OVS-1", "Golden image, setup, diagnostic, backup"),
    ]
    for i, row_data in enumerate(tools, r + 1):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val).border = THIN_BORDER
            ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True)

    auto_width(ws, max_width=40)


def build_epics_lbc(wb):
    """Sheet 5: Epics & LBC — 21 Epics, SUMIF Cost dynamique, WSJF formulas."""
    ws = wb.create_sheet("Epics & LBC")
    ws["A1"] = "\u2190 Dashboard"
    ws["A1"].font = LINK_FONT
    ws["B1"] = "\u2190 Dashboard"
    ws["B1"].font = LINK_FONT

    headers = ["Epic ID", "Name", "VS", "Theme", "WSJF", "PI", "Status", "Cost (SP)", "Benefit (\u20ac)", "Business Value", "Time Criticality", "Job Size"]
    set_header_row(ws, 3, headers, fill=BLUE_LIGHT)

    epics = [
        ("E-01", "Portail Sponsor Self-Service", "VS2 - Sponsor to Impression", "TS1 - Mon\u00e9tisation", None, "PI-1", "Backlog", None, None, 5, 8, 1),
        ("E-02", "Rotation Sponsors", "VS2 - Sponsor to Impression", "TS1 - Mon\u00e9tisation", None, "PI-1", "Backlog", None, None, 5, 5, 1),
        ("E-03", "Analytics Sponsors Avanc\u00e9", "VS2 - Sponsor to Impression", "TS1 - Mon\u00e9tisation", None, "PI-1", "Backlog", None, None, 8, 13, 1),
        ("E-04", "Profils Config Match", "VS1 - Club to Screen", "TS2 - Exp\u00e9rience Match", None, "PI-1", "Termin\u00e9", None, None, 3, 5, 1),
        ("E-06", "Onboarding Automatis\u00e9", "VS1 - Club to Screen", "TS3 - Acquisition", None, "PI-1", "Backlog", None, None, 5, 8, 2),
        ("E-07", "R\u00e9silience WiFi V2", "VS1 - Club to Screen", "TS3 - Acquisition", None, "PI-1", "En cours", None, None, 8, 13, 1),
        ("E-08", "Alertes Pr\u00e9dictives", "Transverse", "TS4 - Excellence Ops", None, "PI-1", "Termin\u00e9", None, None, 5, 8, 1),
        ("E-09", "Architecture Audit", "Transverse", "TS4 - Excellence Ops", None, "PI-1", "Termin\u00e9", None, None, 5, 5, 1),
        ("E-10", "Monitoring Fleet", "Transverse", "TS4 - Excellence Ops", None, "PI-1", "En cours", None, None, 8, 13, 2),
        ("E-05", "Motion Design Personnalis\u00e9", "VS2 - Sponsor to Impression", "TS2 - Exp\u00e9rience Match", None, "PI-2", "Backlog", None, None, 5, 5, 1),
        ("E-11", "R\u00e9gie Publicitaire R\u00e9gionale", "VS2 - Sponsor to Impression", "TS1 - Mon\u00e9tisation", None, "PI-2", "Backlog", None, None, 3, 5, 1),
        ("E-15", "Score en Live Phase 2 (API F\u00e9d\u00e9rations)", "VS1 - Club to Screen", "TS2 - Exp\u00e9rience Match", None, "PI-2", "Backlog", None, None, 3, 5, 1),
        ("E-16", "Rapports Email Automatiques", "Transverse", "TS4 - Excellence Ops", None, "PI-2", "Backlog", None, None, 2, 5, 2),
        ("E-17", "A/B Testing Cr\u00e9as Sponsors", "VS2 - Sponsor to Impression", "TS1 - Mon\u00e9tisation", None, "PI-2", "Backlog", None, None, 2, 3, 1),
        ("E-12", "Multi-\u00c9crans Synchronis\u00e9s", "VS1 - Club to Screen", "TS2 - Exp\u00e9rience Match", None, "PI-3", "Backlog", None, None, 2, 3, 1),
        ("E-13", "Marque Blanche Club", "VS1 - Club to Screen", "TS2 - Exp\u00e9rience Match", None, "PI-3", "Backlog", None, None, 3, 5, 1),
        ("E-14", "Fonds de Solidarit\u00e9 Sport", "Transverse", "TS3 - Acquisition", None, "PI-3", "Backlog", None, None, 3, 5, 2),
        ("E-18", "Int\u00e9grations Billetterie", "VS1 - Club to Screen", "TS2 - Exp\u00e9rience Match", None, "PI-3", "Backlog", None, None, 2, 3, 1),
        ("E-19", "Capteurs Pr\u00e9sence Hardware", "VS1 - Club to Screen", "TS2 - Exp\u00e9rience Match", None, "PI-3", "Backlog", None, None, 2, 2, 3),
        ("E-20", "Analytics Pr\u00e9dictives ML", "Transverse", "TS4 - Excellence Ops", None, "PI-3", "Backlog", None, None, 2, 3, 1),
        ("E-21", "API Partenaires OAuth", "Transverse", "TS1 - Mon\u00e9tisation", None, "PI-3", "Backlog", None, None, 2, 3, 1),
        ("E-22", "Contenus Diff\u00e9renci\u00e9s TV+LED", "VS1 - Club to Screen", "TS2 - Exp\u00e9rience Match", None, "PI-2", "Backlog", None, None, 4, 5, 1),
    ]

    for i, epic in enumerate(epics, 4):
        row_data = list(epic)
        # WSJF formula
        row_data[4] = f"=IF(L{i}=0,0,(J{i}+K{i})/L{i})"
        # Cost = SUMIF from Features & US
        row_data[7] = f"=SUMIF('Features & US'!A:A,A{i},'Features & US'!E:E)"
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER
            if j == 7:
                if val == "Termin\u00e9":
                    cell.fill = GREEN_LIGHT
                elif val == "En cours":
                    cell.fill = ORANGE_LIGHT

    auto_width(ws)


def build_features_us(wb):
    """Sheet 6: Features & US — grouped by PI, COUNTIF/SUMIF → User Stories."""
    ws = wb.create_sheet("Features & US")
    ws["A1"] = "\u2190 Dashboard"
    ws["A1"].font = LINK_FONT
    ws["B1"] = "\u2190 Dashboard"
    ws["B1"].font = LINK_FONT

    ws["A3"] = "FEATURE BACKLOG"
    ws["A3"].font = HEADER_FONT
    headers = ["Epic", "Feature ID", "Feature", "US Count", "SP", "Priority", "Statut", "Notes", "PI", "Sprint"]
    set_header_row(ws, 4, headers, fill=BLUE_LIGHT)

    # PI-1 separator
    ws.cell(row=5, column=1, value="PI-1 (Jan-Mar 2025)").font = HEADER_FONT

    # PI-1 features (rows 6-29)
    features_pi1 = [
        ("E-01", "F-01.1", "Inscription et profil sponsor", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S6"),
        ("E-01", "F-01.2", "Upload vid\u00e9o sponsor", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S5"),
        ("E-01", "F-01.3", "Validation admin des spots", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S5"),
        ("E-02", "F-02.1", "Algorithme de rotation \u00e9quitable", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S5"),
        ("E-03", "F-03.1", "Dashboard impressions sponsor", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S6"),
        ("E-03", "F-03.2", "Export rapport PDF/CSV", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S6"),
        ("E-04", "F-04.1", "Cr\u00e9ation de profils pr\u00e9d\u00e9finis", 1, "Termin\u00e9", "PI-1 (Jan-Mar 2025)", "S2"),
        ("E-04", "F-04.2", "Switch depuis la t\u00e9l\u00e9commande", 1, "Termin\u00e9", "PI-1 (Jan-Mar 2025)", "S3"),
        ("E-05", "F-05.1", "Biblioth\u00e8que de templates motion design", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S6"),
        ("E-06", "F-06.1", "Auto-provisioning Pi", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S5"),
        ("E-06", "F-06.2", "Wizard de configuration club", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S6"),
        ("E-07", "F-07.1", "Cache local \u00e9tendu (48h)", 1, "Termin\u00e9", "PI-1 (Jan-Mar 2025)", "S4"),
        ("E-07", "F-07.2", "Monitoring signal WiFi", 1, "Termin\u00e9", "PI-1 (Jan-Mar 2025)", "S2"),
        ("E-08", "F-08.1", "R\u00e8gles d'alertes pr\u00e9dictives", 1, "Termin\u00e9", "PI-1 (Jan-Mar 2025)", "S2"),
        ("E-08", "F-08.2", "Dashboard tendances", 1, "Termin\u00e9", "PI-1 (Jan-Mar 2025)", "S3"),
        ("E-09", "F-09.1", "Migration controllers vers repository pattern", 1, "Termin\u00e9", "PI-1 (Jan-Mar 2025)", "S4"),
        ("E-09", "F-09.2", "Audit s\u00e9curit\u00e9 et performance", 1, "Termin\u00e9", "PI-1 (Jan-Mar 2025)", "S1"),
        ("E-10", "F-10.1", "Carte de la flotte (Leaflet)", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S6"),
        ("E-10", "F-10.2", "M\u00e9triques agr\u00e9g\u00e9es flotte", 1, "Termin\u00e9", "PI-1 (Jan-Mar 2025)", "S1"),
        ("E-11", "F-11.1", "Portail annonceur r\u00e9gional", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S6"),
        ("E-11", "F-11.2", "Reporting consolid\u00e9 r\u00e9gie", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S5"),
        ("E-12", "F-12.1", "Synchronisation master/slave", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S5"),
        ("E-13", "F-13.1", "Th\u00e9matisation par club", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S5"),
        ("E-16", "F-16.1", "Envoi automatique mensuel", 1, "Backlog", "PI-1 (Jan-Mar 2025)", "S6"),
    ]

    r = 6
    for feat in features_pi1:
        epic, fid, name, priority, statut, pi, sprint = feat
        ws.cell(row=r, column=1, value=epic).border = THIN_BORDER
        ws.cell(row=r, column=2, value=fid).border = THIN_BORDER
        ws.cell(row=r, column=3, value=name).border = THIN_BORDER
        # US Count = COUNTIF from User Stories
        ws.cell(row=r, column=4, value=f"=COUNTIF('User Stories'!B:B,B{r})").border = THIN_BORDER
        # SP = SUMIF from User Stories
        ws.cell(row=r, column=5, value=f"=SUMIF('User Stories'!B:B,B{r},'User Stories'!E:E)").border = THIN_BORDER
        ws.cell(row=r, column=6, value=priority).border = THIN_BORDER
        cell_status = ws.cell(row=r, column=7, value=statut)
        cell_status.border = THIN_BORDER
        if statut == "Termin\u00e9":
            cell_status.fill = GREEN_LIGHT
        ws.cell(row=r, column=8, value="").border = THIN_BORDER
        ws.cell(row=r, column=9, value=pi).border = THIN_BORDER
        ws.cell(row=r, column=10, value=sprint).border = THIN_BORDER
        r += 1

    # PI-2 separator
    ws.cell(row=r, column=1, value="PI-2 (Avr-Jun 2025)").font = HEADER_FONT
    r += 1

    features_pi2 = [
        ("E-02", "F-02.2", "Configuration rotation par gymnase", 2, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-03", "F-03.3", "Heatmap de diffusion", 2, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-05", "F-05.2", "Upload d'animations custom (Lottie/MP4)", 2, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-07", "F-07.3", "Support cl\u00e9 USB WiFi externe", 2, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-14", "F-14.1", "Gestion du fonds", 2, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-15", "F-15.1", "Int\u00e9gration API f\u00e9d\u00e9rations sportives", 2, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-17", "F-17.1", "Campagnes A/B Test", 2, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-18", "F-18.1", "Audience r\u00e9elle via billetterie", 2, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-19", "F-19.1", "Comptage spectateurs automatique", 3, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-20", "F-20.1", "Pr\u00e9dictions engagement et uptime", 3, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-21", "F-21.1", "API OAuth 2.0 pour partenaires", 3, "Backlog", "PI-2 (Avr-Jun 2025)", ""),
        ("E-22", "F-22.0", "Enabler \u2014 Validation hardware dual HDMI (spike)", 1, "Backlog", "PI-2 (Avr-Jun 2026)", ""),
        ("E-22", "F-22.1", "Dual Kiosk HDMI natif", 2, "Backlog", "PI-2 (Avr-Jun 2026)", ""),
        ("E-22", "F-22.2", "R\u00e9actions diff\u00e9renci\u00e9es TV vs LED", 2, "Backlog", "PI-2 (Avr-Jun 2026)", ""),
        ("E-22", "F-22.3", "Variantes vid\u00e9o par type d'\u00e9cran", 2, "Backlog", "PI-2 (Avr-Jun 2026)", ""),
    ]

    pi2_start = r
    for feat in features_pi2:
        epic, fid, name, priority, statut, pi, sprint = feat
        ws.cell(row=r, column=1, value=epic).border = THIN_BORDER
        ws.cell(row=r, column=2, value=fid).border = THIN_BORDER
        ws.cell(row=r, column=3, value=name).border = THIN_BORDER
        ws.cell(row=r, column=4, value=f"=COUNTIF('User Stories'!B:B,B{r})").border = THIN_BORDER
        ws.cell(row=r, column=5, value=f"=SUMIF('User Stories'!B:B,B{r},'User Stories'!E:E)").border = THIN_BORDER
        ws.cell(row=r, column=6, value=priority).border = THIN_BORDER
        ws.cell(row=r, column=7, value=statut).border = THIN_BORDER
        ws.cell(row=r, column=8, value="").border = THIN_BORDER
        ws.cell(row=r, column=9, value=pi).border = THIN_BORDER
        ws.cell(row=r, column=10, value=sprint).border = THIN_BORDER
        r += 1

    pi2_end = r - 1

    # Summary
    r += 1
    ws.cell(row=r, column=1, value="SUMMARY").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["PI", "Features", "Total SP", "Termin\u00e9", "Backlog"])
    r += 1
    write_row(ws, r, [
        "PI-1 (Jan-Mar 2025)",
        "=COUNTIF(I6:I29,\"PI-1*\")",
        "=SUMIF(I6:I29,\"PI-1*\",E6:E29)",
        "=COUNTIFS(I6:I29,\"PI-1*\",G6:G29,\"Termin\u00e9\")",
        "=COUNTIFS(I6:I29,\"PI-1*\",G6:G29,\"Backlog\")",
    ])
    r += 1
    write_row(ws, r, [
        "PI-2 (Avr-Jun 2025)",
        f"=COUNTIF(I{pi2_start}:I{pi2_end},\"PI-2*\")",
        f"=SUMIF(I{pi2_start}:I{pi2_end},\"PI-2*\",E{pi2_start}:E{pi2_end})",
        f"=COUNTIFS(I{pi2_start}:I{pi2_end},\"PI-2*\",G{pi2_start}:G{pi2_end},\"Termin\u00e9\")",
        f"=COUNTIFS(I{pi2_start}:I{pi2_end},\"PI-2*\",G{pi2_start}:G{pi2_end},\"Backlog\")",
    ])
    r += 1
    write_row(ws, r, ["PI-3 (Jul-Sep 2025)", 0, 0, 0, 0])

    auto_width(ws)


def build_pi_objectives(wb):
    """Sheet 7: PI Objectives — PI-1/2 + aspirationnel, 7 cols, r\u00e9sum\u00e9 SUMIFS."""
    ws = wb.create_sheet("PI Objectives")
    ws["A1"] = "\u2190 Dashboard"
    ws["A1"].font = LINK_FONT
    ws["B1"] = "\u2190 Dashboard"
    ws["B1"].font = LINK_FONT

    ws["A3"] = "OBJECTIFS PI"
    ws["A3"].font = HEADER_FONT
    headers = ["Objectif", "Type", "Description", "Valeur M\u00e9tier", "Statut", "Notes / Features", "PI"]
    set_header_row(ws, 4, headers, fill=GREEN_LIGHT)

    # PI-1 separator
    ws.cell(row=5, column=1, value="PI-1 (Jan-Mar 2025)").font = HEADER_FONT

    objectives_pi1 = [
        ("O-PI1-1", "Engag\u00e9", "Lancer le portail sponsor self-service", 9, "\u00c0 faire", "E-01: F-01.1, F-01.2, F-01.3", "PI-1"),
        ("O-PI1-2", "Engag\u00e9", "Livrer les analytics sponsors avec rapport PDF", 10, "\u00c0 faire", "E-03: F-03.1, F-03.2, F-03.3", "PI-1"),
        ("O-PI1-3", "Engag\u00e9", "Impl\u00e9menter la rotation sponsor \u00e9quitable", 8, "\u00c0 faire", "E-02: F-02.1, F-02.2", "PI-1"),
        ("O-PI1-4", "Engag\u00e9", "Cr\u00e9er le wizard onboarding club", 10, "\u00c0 faire", "E-06: F-06.1, F-06.2", "PI-1"),
        ("O-PI1-5", "\u00c9tendu", "Carte de la flotte Leaflet", 4, "\u00c0 faire", "E-10: F-10.1", "PI-1"),
    ]
    r = 6
    for obj in objectives_pi1:
        for j, val in enumerate(obj, 1):
            ws.cell(row=r, column=j, value=val).border = THIN_BORDER
        r += 1

    # PI-2 separator
    ws.cell(row=r, column=1, value="PI-2 (Avr-Jun 2025)").font = HEADER_FONT
    r += 1

    objectives_pi2 = [
        ("O-PI2-1", "\u00c9tendu", "Support cl\u00e9 USB WiFi externe", 3, "\u00c0 faire", "E-07: F-07.3", "PI-2"),
        ("O-PI2-2", "\u00c9tendu", "Diffusion dynamique de publicit\u00e9s", 8, "En cours", "E-11: F-11.1, F-11.2", "PI-2"),
        ("O-PI2-3", "\u00c9tendu", "Plateforme d'email marketing", 7, "\u00c0 faire", "E-15: F-15.1", "PI-2"),
        ("O-PI2-4", "\u00c9tendu", "Scoring influenceurs pr\u00eat", 6, "\u00c0 faire", "E-16: F-16.1 + E-17: F-17.1", "PI-2"),
    ]
    for obj in objectives_pi2:
        for j, val in enumerate(obj, 1):
            ws.cell(row=r, column=j, value=val).border = THIN_BORDER
        r += 1

    # Aspirationnel
    ws.cell(row=r, column=1, value="ASPIRATIONNEL (Non planifi\u00e9s)").font = HEADER_FONT
    r += 1
    aspirationnel = [
        ("O-STR-1", "Aspirationnel", "Analytics avanc\u00e9s ML", 4, "\u00c0 faire", "\u2014", "\u2014"),
        ("O-STR-2", "Aspirationnel", "\u00c9bauche expansion internationale", 3, "\u00c0 faire", "\u2014", "\u2014"),
    ]
    for obj in aspirationnel:
        for j, val in enumerate(obj, 1):
            ws.cell(row=r, column=j, value=val).border = THIN_BORDER
        r += 1

    # Résumé avec formules SUMIFS
    r += 1
    ws.cell(row=r, column=1, value="R\u00c9SUM\u00c9").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["PI", "BV Engag\u00e9", "BV \u00c9tendu", "BV Aspirationnel", "BV Total", "Obj. Termin\u00e9s", "Pr\u00e9dictabilit\u00e9 %"])
    r += 1
    write_row(ws, r, [
        "PI-1",
        "=SUMIFS(D6:D18,B6:B18,\"Engag\u00e9\",G6:G18,\"PI-1\")",
        "=SUMIFS(D6:D18,B6:B18,\"\u00c9tendu\",G6:G18,\"PI-1\")",
        "=SUMIFS(D17:D18,B17:B18,\"Aspirationnel\")",
        "=SUM(B{0}:D{0})".format(r),
        "=COUNTIFS(G6:G10,\"PI-1\",E6:E10,\"Termin\u00e9\")",
        "=IFERROR(F{0}/COUNTA(A6:A10),0)".format(r),
    ])
    r_pi1 = r
    r += 1
    write_row(ws, r, [
        "PI-2",
        "=SUMIFS(D6:D18,B6:B18,\"Engag\u00e9\",G6:G18,\"PI-2\")",
        "=SUMIFS(D6:D18,B6:B18,\"\u00c9tendu\",G6:G18,\"PI-2\")",
        0,
        "=SUM(B{0}:D{0})".format(r),
        "=COUNTIFS(G12:G15,\"PI-2\",E12:E15,\"Termin\u00e9\")",
        "=IFERROR(F{0}/COUNTA(A12:A15),0)".format(r),
    ])
    r_pi2 = r
    r += 1
    write_row(ws, r, [
        "Total",
        f"=SUM(B{r_pi1}:B{r_pi2})",
        f"=SUM(C{r_pi1}:C{r_pi2})",
        f"=SUM(D{r_pi1}:D{r_pi2})",
        f"=SUM(E{r_pi1}:E{r_pi2})",
        f"=SUM(F{r_pi1}:F{r_pi2})",
        f"=IFERROR(F{r}/COUNTA(A6:A18),0)",
    ])

    auto_width(ws)


def build_sprint_tracker(wb):
    """Sheet 8: Sprint Tracker — S1-S12 (PI-1 + PI-2), r\u00e9sum\u00e9, capacit\u00e9 \u00e9quipe."""
    ws = wb.create_sheet("Sprint Tracker")
    ws["A1"] = "\u2190 Dashboard"
    ws["A1"].font = LINK_FONT
    ws["B1"] = "\u2190 Dashboard"
    ws["B1"].font = LINK_FONT

    ws["A3"] = "SUIVI DES SPRINTS"
    ws["A3"].font = HEADER_FONT
    headers = ["Sprint", "PI", "SP Planifi\u00e9s", "SP R\u00e9alis\u00e9s", "Bugs", "V\u00e9locit\u00e9", "Notes"]
    set_header_row(ws, 4, headers, fill=BLUE_LIGHT)

    # PI-1 separator
    ws.cell(row=5, column=1, value="PI-1 (Jan-Mar 2025)").font = HEADER_FONT

    sprints_pi1 = [
        ("S1", "PI-1", 24, 3, "Premier sprint, onboarding"),
        ("S2", "PI-1", 27, 0, "V\u00e9locit\u00e9 maximale"),
        ("S3", "PI-1", 25, 0, "Performance stable"),
        ("S4", "PI-1", 25, 0, "Performance stable"),
        ("S5", "PI-1", None, None, ""),
        ("S6", "PI-1", None, None, ""),
    ]
    r = 6
    for sprint_name, pi, completed, bugs, notes in sprints_pi1:
        sprint_num = sprint_name.replace("S", "")
        ws.cell(row=r, column=1, value=sprint_name).border = THIN_BORDER
        ws.cell(row=r, column=2, value=pi).border = THIN_BORDER
        ws.cell(row=r, column=3, value=f"=SUMIF('Features & US'!J:J,\"{sprint_name}\",'Features & US'!E:E)").border = THIN_BORDER
        ws.cell(row=r, column=4, value=completed).border = THIN_BORDER
        ws.cell(row=r, column=5, value=bugs).border = THIN_BORDER
        ws.cell(row=r, column=6, value=f"=IF(D{r}=\"\",\"\",D{r})").border = THIN_BORDER
        ws.cell(row=r, column=7, value=notes).border = THIN_BORDER
        r += 1

    # PI-2 separator
    ws.cell(row=r, column=1, value="PI-2 (Avr-Jun 2025)").font = HEADER_FONT
    r += 1

    sprints_pi2 = [
        ("S7", "PI-2"), ("S8", "PI-2"), ("S9", "PI-2"),
        ("S10", "PI-2"), ("S11", "PI-2"), ("S12", "PI-2"),
    ]
    pi2_start = r
    for sprint_name, pi in sprints_pi2:
        ws.cell(row=r, column=1, value=sprint_name).border = THIN_BORDER
        ws.cell(row=r, column=2, value=pi).border = THIN_BORDER
        ws.cell(row=r, column=3, value=f"=SUMIF('Features & US'!J:J,\"{sprint_name}\",'Features & US'!E:E)").border = THIN_BORDER
        ws.cell(row=r, column=4, value="").border = THIN_BORDER
        ws.cell(row=r, column=5, value="").border = THIN_BORDER
        ws.cell(row=r, column=6, value=f"=IF(D{r}=\"\",\"\",D{r})").border = THIN_BORDER
        ws.cell(row=r, column=7, value="").border = THIN_BORDER
        r += 1
    pi2_end = r - 1

    # Résumé
    r += 1
    ws.cell(row=r, column=1, value="R\u00c9SUM\u00c9").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["PI", "Sprints termin\u00e9s", "SP Planifi\u00e9s", "SP R\u00e9alis\u00e9s", "Total Bugs", "V\u00e9locit\u00e9 moy.", "Engagement %"])
    r += 1
    r_pi1_sum = r
    write_row(ws, r, [
        "PI-1",
        "=COUNTA(D6:D11)",
        "=SUM(C6:C11)",
        "=SUM(D6:D11)",
        "=SUM(E6:E11)",
        "=IFERROR(AVERAGE(F6:F11),0)",
        "=IFERROR(D{0}/C{0},0)".format(r),
    ])
    r += 1
    r_pi2_sum = r
    write_row(ws, r, [
        "PI-2",
        f"=COUNTA(D{pi2_start}:D{pi2_end})",
        f"=SUM(C{pi2_start}:C{pi2_end})",
        f"=SUM(D{pi2_start}:D{pi2_end})",
        f"=SUM(E{pi2_start}:E{pi2_end})",
        f"=IFERROR(AVERAGE(F{pi2_start}:F{pi2_end}),0)",
        "=IFERROR(D{0}/C{0},0)".format(r),
    ])
    r += 1
    write_row(ws, r, [
        "Total",
        f"=SUM(B{r_pi1_sum}:B{r_pi2_sum})",
        f"=SUM(C{r_pi1_sum}:C{r_pi2_sum})",
        f"=SUM(D{r_pi1_sum}:D{r_pi2_sum})",
        f"=SUM(E{r_pi1_sum}:E{r_pi2_sum})",
        f"=IFERROR(AVERAGE(F{r_pi1_sum}:F{r_pi2_sum}),0)",
        "=IFERROR(D{0}/C{0},0)".format(r),
    ])

    # Capacité équipe
    r += 2
    ws.cell(row=r, column=1, value="CAPACIT\u00c9 DE L'\u00c9QUIPE").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["Membre", "R\u00f4le", "Capacit\u00e9 (SP/sprint)", "Sp\u00e9cialit\u00e9"], fill=GREY_LIGHT)
    r += 1
    team = [
        ("Dev Lead", "", 8, "Backend / Architecture"),
        ("Frontend Dev", "", 7, "UI / Angular"),
        ("Full Stack Dev", "", 8, "Fonctionnalit\u00e9s"),
        ("QA Engineer", "", 4, "Tests / Qualit\u00e9"),
    ]
    team_start = r
    for member, role, cap, spec in team:
        write_row(ws, r, [member, role, cap, spec])
        r += 1
    team_end = r - 1
    write_row(ws, r, ["Total", "", f"=SUM(C{team_start}:C{team_end})"])

    auto_width(ws)


def build_roam(wb):
    """Sheet 9: ROAM — 8 colonnes, r\u00e9sum\u00e9 COUNTIF."""
    ws = wb.create_sheet("ROAM")
    ws["A1"] = "\u2190 Dashboard"
    ws["A1"].font = LINK_FONT
    ws["B1"] = "\u2190 Dashboard"
    ws["B1"].font = LINK_FONT

    ws["A3"] = "REGISTRE ROAM"
    ws["A3"].font = HEADER_FONT
    headers = ["ID Risque", "Description", "Statut ROAM", "Probabilit\u00e9", "Impact", "Responsable", "\u00c9ch\u00e9ance", "Action / Att\u00e9nuation"]
    set_header_row(ws, 4, headers, fill=RED_LIGHT)

    risks = [
        ("R-01", "Capacit\u00e9 solo-dev insuffisante pour 81 SP engag\u00e9s", "Accept\u00e9", "Haute", "Moyen", "Gwenvael", "", ""),
        ("R-02", "WiFi gymnase instable pendant les tests", "Att\u00e9nu\u00e9", "Haute", "Moyen", "Gwenvael", "", ""),
        ("R-03", "Aucun sponsor inscrit pour valider le portail", "Pris en charge", "Moyenne", "\u00c9lev\u00e9", "Gwenvael + Gabin", "", ""),
        ("R-04", "D\u00e9pendance Supabase pour le scaling", "Accept\u00e9", "Basse", "\u00c9lev\u00e9", "Gwenvael", "", ""),
        ("R-05", "S\u00e9curit\u00e9 des api_keys Pi", "Att\u00e9nu\u00e9", "Basse", "Critique", "Gwenvael", "", ""),
        ("R-06", "Retard onboarding automatis\u00e9 bloque le scaling", "Pris en charge", "Moyenne", "Critique", "Gwenvael", "", ""),
        ("R-07", "H\u00e9bergement FTP Hostinger comme point de d\u00e9faillance", "Accept\u00e9", "Basse", "\u00c9lev\u00e9", "Gwenvael", "", ""),
        ("R-08", "Absence de tests E2E sur le parcours sponsor", "Pris en charge", "Haute", "Moyen", "Gwenvael", "", ""),
    ]

    r = 5
    for risk in risks:
        for j, val in enumerate(risk, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.border = THIN_BORDER
        r += 1
    risk_end = r - 1

    # Résumé ROAM avec formules COUNTIF
    r += 1
    ws.cell(row=r, column=1, value="R\u00c9SUM\u00c9 ROAM").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["Statut ROAM", "Nombre", "% du Total", "Couleur", "", "Niveau Impact", "Nombre", "% du Total"])
    r += 1
    r_start = r

    roam_summary = [
        ("R\u00e9solu", f'=COUNTIF(C5:C{risk_end},"R\u00e9solu")', None, "\u2588 Vert", "", "Critique", f'=COUNTIF(E5:E{risk_end},"Critique")', None),
        ("Pris en charge", f'=COUNTIF(C5:C{risk_end},"Pris en charge")', None, "\u2588 Orange", "", "\u00c9lev\u00e9", f'=COUNTIF(E5:E{risk_end},"\u00c9lev\u00e9")', None),
        ("Accept\u00e9", f'=COUNTIF(C5:C{risk_end},"Accept\u00e9")', None, "\u2588 Jaune", "", "Moyen", f'=COUNTIF(E5:E{risk_end},"Moyen")', None),
        ("Att\u00e9nu\u00e9", f'=COUNTIF(C5:C{risk_end},"Att\u00e9nu\u00e9")', None, "\u2588 Bleu", "", "TOTAL", None, None),
    ]

    for i, (label, count_formula, pct, color, _, imp_label, imp_count, imp_pct) in enumerate(roam_summary):
        row = r + i
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=count_formula).border = THIN_BORDER
        total_row = r_start + 4
        ws.cell(row=row, column=3, value=f"=IFERROR(B{row}/B{total_row},0)").border = THIN_BORDER
        ws.cell(row=row, column=4, value=color).border = THIN_BORDER
        ws.cell(row=row, column=6, value=imp_label).border = THIN_BORDER
        if imp_count:
            ws.cell(row=row, column=7, value=imp_count).border = THIN_BORDER
        if i < 3:
            imp_total_row = r_start + 3
            ws.cell(row=row, column=8, value=f"=IFERROR(G{row}/G{imp_total_row},0)").border = THIN_BORDER

    # Totals
    total_row = r_start + 4
    ws.cell(row=total_row, column=1, value="TOTAL").border = THIN_BORDER
    ws.cell(row=total_row, column=2, value=f"=SUM(B{r_start}:B{r_start+3})").border = THIN_BORDER
    ws.cell(row=total_row, column=3, value=f"=SUM(C{r_start}:C{r_start+3})").border = THIN_BORDER

    # Impact totals
    imp_total = r_start + 3  # TOTAL row for impact
    ws.cell(row=imp_total, column=7, value=f"=SUM(G{r_start}:G{r_start+2})").border = THIN_BORDER
    ws.cell(row=imp_total, column=8, value=f"=SUM(H{r_start}:H{r_start+2})").border = THIN_BORDER

    auto_width(ws)


def build_flow_metrics(wb):
    """Sheet 10: Flow Metrics — m\u00e9triques VS + allocation avec barres REPT."""
    ws = wb.create_sheet("Flow Metrics")
    ws["A1"] = "\u2190 Dashboard"
    ws["A1"].font = LINK_FONT
    ws["B1"] = "\u2190 Dashboard"
    ws["B1"].font = LINK_FONT

    ws["A3"] = "M\u00c9TRIQUES DE FLUX PAR VALUE STREAM"
    ws["A3"].font = HEADER_FONT
    headers = ["Value Stream", "Limite TEP", "D\u00e9bit (US/sem.)", "Temps de Cycle (jours)", "Sant\u00e9 du Flux"]
    set_header_row(ws, 4, headers, fill=PURPLE_LIGHT)

    vs_metrics = [
        "OVS-1 : Club to Screen",
        "OVS-2 : Sponsor Impression",
        "DVS-1 : Platform",
        "Support / Dette technique",
        "Int\u00e9gration / Infra",
        "Apprentissage / Innovation",
    ]
    r = 5
    for vs in vs_metrics:
        ws.cell(row=r, column=1, value=vs).border = THIN_BORDER
        for j in range(2, 6):
            ws.cell(row=r, column=j, value="").border = THIN_BORDER
        r += 1

    ws.cell(row=r, column=1, value="TOTAL / MOYENNE").border = THIN_BORDER
    ws.cell(row=r, column=2, value="=SUM(B5:B10)").border = THIN_BORDER
    ws.cell(row=r, column=3, value="=AVERAGE(C5:C10)").border = THIN_BORDER
    ws.cell(row=r, column=4, value="=AVERAGE(D5:D10)").border = THIN_BORDER
    r += 2

    # Allocation par VS avec barres REPT
    ws.cell(row=r, column=1, value="ALLOCATION PAR VALUE STREAM").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["Value Stream", "Allocation %", "Barre visuelle"], fill=GREY_LIGHT)
    r += 1
    alloc_start = r
    alloc_vs = ["Club to Screen", "Sponsor Impression", "Platform Core", "Support / Dette", "Apprentissage"]
    for vs in alloc_vs:
        ws.cell(row=r, column=1, value=vs).border = THIN_BORDER
        ws.cell(row=r, column=2, value="").border = THIN_BORDER
        ws.cell(row=r, column=3, value=f'=REPT("\u2588",B{r}/5)').border = THIN_BORDER
        r += 1
    alloc_end = r - 1
    ws.cell(row=r, column=1, value="TOTAL").border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"=SUM(B{alloc_start}:B{alloc_end})").border = THIN_BORDER

    auto_width(ws)


def build_implemented_backlog(wb):
    """Sheet 11: Implemented Backlog — from CSV or .md, + stats + domain summary COUNTIF."""
    ws = wb.create_sheet("Implemented Backlog")
    ws["A1"] = "\u2190 Dashboard"
    ws["A1"].font = LINK_FONT
    ws["B1"] = "\u2190 Dashboard"
    ws["B1"].font = LINK_FONT

    csv_path = SAFE_DIR / "notion-import" / "implemented-backlog-import.csv"
    if csv_path.exists():
        ws["A2"] = f"184 features \u2014 13 domaines \u2014 Source: {csv_path.name}"
    else:
        ws["A2"] = "Source: IMPLEMENTED-BACKLOG.md"

    headers_imp = ["ID", "Feature", "Status", "Version", "Fichier", "Domaine"]
    set_header_row(ws, 3, headers_imp, fill=GREY_LIGHT)

    row_idx = 4
    current_domain = ""

    if csv_path.exists():
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)  # skip
            for row_data in reader:
                for j, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=j, value=val)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(wrap_text=True)
                row_idx += 1
    else:
        md = read_md("IMPLEMENTED-BACKLOG.md")
        for line in md.split("\n"):
            line = line.strip()
            if line.startswith("## ") and not line.startswith("## R\u00e9sum\u00e9"):
                current_domain = re.sub(r"^## \d+\. ", "", line)
                ws.cell(row=row_idx, column=1, value=current_domain).font = HEADER_FONT
                row_idx += 1
            elif line.startswith("|") and "IMP-" in line:
                cells = [c.strip() for c in line.split("|")[1:-1]]
                if len(cells) >= 4:
                    ws.cell(row=row_idx, column=1, value=cells[0]).border = THIN_BORDER
                    ws.cell(row=row_idx, column=2, value=cells[1] if len(cells) > 1 else "").border = THIN_BORDER
                    ws.cell(row=row_idx, column=3, value=cells[2] if len(cells) > 2 else "").border = THIN_BORDER
                    ws.cell(row=row_idx, column=4, value=cells[3] if len(cells) > 3 else "").border = THIN_BORDER
                    ws.cell(row=row_idx, column=5, value=cells[4] if len(cells) > 4 else "").border = THIN_BORDER
                    ws.cell(row=row_idx, column=6, value=current_domain).border = THIN_BORDER
                    row_idx += 1

    # Statistiques Produit
    r = row_idx + 1
    ws.cell(row=r, column=1, value="Statistiques Produit").font = HEADER_FONT
    r += 1
    stats = [
        ("Features impl\u00e9ment\u00e9es", "184"),
        ("Domaines fonctionnels", ""),
        ("Controllers API", "38"),
        ("Services m\u00e9tier", "37"),
        ("Repositories", "24"),
        ("Versions released", "30+ (v1.0 \u2192 v3.60.0)"),
        ("Tests (total)", "2 235"),
    ]
    for label, val in stats:
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        ws.cell(row=r, column=2, value=val).border = THIN_BORDER
        r += 1

    # Navigation
    r += 1
    ws.cell(row=r, column=1, value="\u2190 Dashboard").font = LINK_FONT
    ws.cell(row=r, column=2, value="\u2190 Features & US").font = LINK_FONT
    r += 2

    # Domain Summary with COUNTIF
    ws.cell(row=r, column=1, value="DOMAIN SUMMARY").font = HEADER_FONT
    r += 1
    set_header_row(ws, r, ["Domaine", "Nombre"])
    r += 1
    domains = [
        "Authentification & S\u00e9curit\u00e9",
        "Clubs & Gestion des sites",
        "Diffusion & Playlists",
        "Content Management",
        "Analytics & Reporting",
        "Sponsors & R\u00e9gie",
        "Mobile & T\u00e9l\u00e9commande",
        "Infrastructure & Monitoring",
        "Admin & Configuration",
        "API & Int\u00e9grations",
        "Notifications",
        "Performance & Scalabilit\u00e9",
        "DevOps & D\u00e9ploiement",
    ]
    domain_start = r
    for domain in domains:
        ws.cell(row=r, column=1, value=domain).border = THIN_BORDER
        ws.cell(row=r, column=2, value=f'=COUNTIF(F:F,"{domain}")').border = THIN_BORDER
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"=SUM(B{domain_start}:B{r-1})").border = THIN_BORDER

    auto_width(ws, max_width=60)


def build_user_stories(wb):
    """Sheet 12: User Stories — backlog header (to be filled)."""
    ws = wb.create_sheet("User Stories")
    ws["A1"] = "\u2190 Dashboard"
    ws["A1"].font = LINK_FONT
    ws["B1"] = "\u2190 Features & US"
    ws["B1"].font = LINK_FONT

    ws["A3"] = "USER STORIES BACKLOG"
    ws["A3"].font = HEADER_FONT
    headers = ["US ID", "Feature ID", "Epic ID", "User Story", "Story Points", "Statut", "Sprint", "Crit\u00e8res d'acceptation"]
    set_header_row(ws, 4, headers, fill=BLUE_LIGHT)

    auto_width(ws)


def build_chart_data(wb):
    """Sheet 13: _ChartData — data for Excel charts."""
    ws = wb.create_sheet("_ChartData")

    # Epics status
    ws["A1"] = "Statut"
    ws["B1"] = "Epics"
    ws["A1"].font = HEADER_FONT
    ws["B1"].font = HEADER_FONT
    ws["A2"] = "En cours"
    ws["B2"] = "=COUNTIF('Epics & LBC'!G:G,\"En cours\")"
    ws["A3"] = "Termin\u00e9"
    ws["B3"] = "=COUNTIF('Epics & LBC'!G:G,\"Termin\u00e9\")"
    ws["A4"] = "Backlog"
    ws["B4"] = "=COUNTIF('Epics & LBC'!G:G,\"Backlog\")"

    # SP by PI
    ws["A6"] = "PI"
    ws["B6"] = "Story Points"
    ws["A6"].font = HEADER_FONT
    ws["B6"].font = HEADER_FONT
    ws["A7"] = "PI-1"
    ws["B7"] = "=SUMIF('Features & US'!I:I,\"PI-1*\",'Features & US'!E:E)"
    ws["A8"] = "PI-2"
    ws["B8"] = "=SUMIF('Features & US'!I:I,\"PI-2*\",'Features & US'!E:E)"
    ws["A9"] = "PI-3"
    ws["B9"] = "=SUMIF('Features & US'!I:I,\"PI-3*\",'Features & US'!E:E)"

    # Sprint velocity
    ws["A11"] = "Sprint"
    ws["B11"] = "Actual SP"
    ws["A11"].font = HEADER_FONT
    ws["B11"].font = HEADER_FONT
    for i, s in enumerate(["S1", "S2", "S3", "S4"], 12):
        ws.cell(row=i, column=1, value=s)
        ws.cell(row=i, column=2, value=f"='Sprint Tracker'!D{i - 12 + 6}")

    # Flow distribution
    ws["A17"] = "Domaine"
    ws["B17"] = "Allocation"
    ws["A17"].font = HEADER_FONT
    ws["B17"].font = HEADER_FONT
    flow = [("Mon\u00e9tisation", 40), ("Exp. Match", 25), ("Acquisition", 20), ("Excellence Ops", 15)]
    for i, (domain, alloc) in enumerate(flow, 18):
        ws.cell(row=i, column=1, value=domain)
        ws.cell(row=i, column=2, value=alloc)

    auto_width(ws)


# ============================================================
# Main
# ============================================================

def main():
    output_path = DEFAULT_OUTPUT
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output_path = Path(sys.argv[idx + 1])

    print("=" * 70)
    print("NEOPRO SAFe Portfolio \u2014 Excel Generator")
    print(f"Source: {SAFE_DIR}/*.md")
    print(f"Output: {output_path}")
    print("=" * 70)

    wb = openpyxl.Workbook()

    steps = [
        ("Dashboard", build_dashboard),
        ("Glossaire", build_glossaire),
        ("Vision & OKR", build_vision_okr),
        ("Value Streams", build_value_streams),
        ("Epics & LBC", build_epics_lbc),
        ("Features & US", build_features_us),
        ("PI Objectives", build_pi_objectives),
        ("Sprint Tracker", build_sprint_tracker),
        ("ROAM", build_roam),
        ("Flow Metrics", build_flow_metrics),
        ("Implemented Backlog", build_implemented_backlog),
        ("User Stories", build_user_stories),
        ("_ChartData", build_chart_data),
    ]

    for i, (name, builder) in enumerate(steps, 1):
        print(f"  [{i:2d}/{len(steps)}] {name}...")
        builder(wb)

    # Set calculation mode
    wb.calculation.calcMode = "auto"

    print(f"\nSaving to {output_path}...")
    wb.save(str(output_path))

    # Verify
    size = output_path.stat().st_size
    print(f"\n\u2705 Generated: {output_path.name} ({size:,} bytes)")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")
    print(f"\n\U0001f4a1 Open in Excel/Google Sheets to recalculate formulas.")
    print("   Or run: python docs/safe/scripts/recalc.py " + str(output_path))


if __name__ == "__main__":
    main()
